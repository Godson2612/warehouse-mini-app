
import os
import io
import csv
import re
import sqlite3
import logging
import threading
import asyncio
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from contextlib import closing

from flask import Flask, request, jsonify, redirect, Response, render_template
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    WebAppInfo,
    InputFile,
    ReplyKeyboardMarkup,
    KeyboardButton,
    BotCommand,
)
from telegram.error import BadRequest
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    CallbackQueryHandler,
    MessageHandler,
    filters,
)

# ============================================================
# CONFIG
# ============================================================
TECH_BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_BOT_TOKEN = os.getenv("ADMIN_BOT_TOKEN", "").strip()
BASE_URL = os.getenv("BASE_URL", "https://warehouse-mini-app.onrender.com").rstrip("/")
APP_HOST = os.getenv("APP_HOST", "0.0.0.0")
APP_PORT = int(os.getenv("PORT", "8080"))
APP_TIMEZONE = os.getenv("APP_TIMEZONE", "America/New_York")
DB_PATH = os.getenv("DB_PATH", "orders.db")
SECRET_KEY = os.getenv("SECRET_KEY", "warehouse-secret-key")
ADMIN_ACCESS_TOKEN = os.getenv("ADMIN_TOKEN", "CHANGE_THIS_ADMIN_TOKEN").strip()
OWNER_ADMIN_ID = os.getenv("OWNER_ADMIN_ID", "").strip()
DEFAULT_ADMIN_IDS = os.getenv("DEFAULT_ADMIN_IDS", "").strip()
RUN_TECH_BOT = os.getenv("RUN_TECH_BOT", "").strip().lower() in {"1", "true", "yes", "on"}
RUN_ADMIN_BOT = os.getenv("RUN_ADMIN_BOT", "").strip().lower() in {"1", "true", "yes", "on"}
SQLITE_TIMEOUT = float(os.getenv("SQLITE_TIMEOUT", "30"))

TZ = ZoneInfo(APP_TIMEZONE)

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("warehouse_app")

app = Flask(__name__)
app.secret_key = SECRET_KEY

tech_bot_app = None
admin_bot_app = None

if TECH_BOT_TOKEN:
    tech_bot_app = Application.builder().token(TECH_BOT_TOKEN).build()

if ADMIN_BOT_TOKEN:
    admin_bot_app = Application.builder().token(ADMIN_BOT_TOKEN).build()

_tech_handlers_registered = False
_admin_handlers_registered = False

# ============================================================
# DATABASE
# ============================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at TEXT NOT NULL,
    created_date TEXT NOT NULL,
    telegram_user_id INTEGER,
    telegram_username TEXT,
    tech_id TEXT NOT NULL,
    bp_number TEXT NOT NULL,

    xb3 INTEGER NOT NULL DEFAULT 0,
    xb6 INTEGER NOT NULL DEFAULT 0,
    xb7 INTEGER NOT NULL DEFAULT 0,
    xb8 INTEGER NOT NULL DEFAULT 0,
    xb10 INTEGER NOT NULL DEFAULT 0,

    xg1 INTEGER NOT NULL DEFAULT 0,
    xg1_4k INTEGER NOT NULL DEFAULT 0,
    xg2 INTEGER NOT NULL DEFAULT 0,
    xid INTEGER NOT NULL DEFAULT 0,
    xi6 INTEGER NOT NULL DEFAULT 0,
    xer10 INTEGER NOT NULL DEFAULT 0,
    onu INTEGER NOT NULL DEFAULT 0,

    screen INTEGER NOT NULL DEFAULT 0,
    battery INTEGER NOT NULL DEFAULT 0,
    sensor INTEGER NOT NULL DEFAULT 0,
    camera INTEGER NOT NULL DEFAULT 0,

    extra_item_name TEXT NOT NULL DEFAULT '',
    extra_item_qty INTEGER NOT NULL DEFAULT 0,

    notes TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    exported_at TEXT NOT NULL DEFAULT ''
);

CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS admin_users (
    telegram_user_id INTEGER PRIMARY KEY,
    telegram_username TEXT NOT NULL DEFAULT '',
    role TEXT NOT NULL DEFAULT 'admin',
    added_at TEXT NOT NULL,
    added_by INTEGER
);
"""

DEFAULT_SETTINGS = {
    "limit_modems_total": "12",
    "limit_dvr_total": "5",
    "limit_xg2": "5",
    "limit_xid": "12",
    "limit_xi6": "12",
    "limit_xer10": "2",
    "limit_onu": "2",
    "limit_screen": "5",
    "limit_battery": "50",
    "limit_sensor": "50",
    "limit_camera": "50",
    "limit_extra_item": "10",
    "technician_message": "",
    "technician_message_active_until": "",
    "owner_admin_id": "",
}

EQUIPMENT_LABELS = {
    "xb3": "XB3",
    "xb6": "XB6",
    "xb7": "XB7",
    "xb8": "XB8",
    "xb10": "XB10",
    "xg1": "XG1",
    "xg1_4k": "XG1 4K",
    "xg2": "XG2",
    "xid": "XID",
    "xi6": "XI6",
    "xer10": "XER10",
    "onu": "ONU",
    "screen": "Screen",
    "battery": "Battery",
    "sensor": "Sensor",
    "camera": "Camera",
    "extra_item_qty": "Additional Item",
}

EQUIPMENT_ORDER = [
    "xb3", "xb6", "xb7", "xb8", "xb10",
    "xg1", "xg1_4k", "xg2", "xid", "xi6",
    "xer10", "onu", "screen", "battery", "sensor", "camera",
    "extra_item_qty",
]

MERGE_SUM_FIELDS = [
    "xb3", "xb6", "xb7", "xb8", "xb10",
    "xg1", "xg1_4k", "xg2", "xid", "xi6",
    "xer10", "onu", "screen", "battery", "sensor", "camera",
    "extra_item_qty",
]


def get_db():
    db_dir = os.path.dirname(DB_PATH)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)

    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=SQLITE_TIMEOUT)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout = 30000")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def column_exists(conn, table_name: str, column_name: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(r["name"] == column_name for r in rows)


def get_setting_from_conn(conn, key: str, default: str = "") -> str:
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting_with_conn(conn, key: str, value: str):
    conn.execute(
        """
        INSERT INTO settings(key, value) VALUES(?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """,
        (key, value),
    )


def init_db():
    with closing(get_db()) as conn:
        conn.executescript(SCHEMA_SQL)

        if not column_exists(conn, "admin_users", "telegram_username"):
            conn.execute("ALTER TABLE admin_users ADD COLUMN telegram_username TEXT NOT NULL DEFAULT ''")
        if not column_exists(conn, "admin_users", "role"):
            conn.execute("ALTER TABLE admin_users ADD COLUMN role TEXT NOT NULL DEFAULT 'admin'")
        if not column_exists(conn, "admin_users", "added_by"):
            conn.execute("ALTER TABLE admin_users ADD COLUMN added_by INTEGER")

        if not column_exists(conn, "orders", "status"):
            conn.execute("ALTER TABLE orders ADD COLUMN status TEXT NOT NULL DEFAULT 'active'")
        if not column_exists(conn, "orders", "exported_at"):
            conn.execute("ALTER TABLE orders ADD COLUMN exported_at TEXT NOT NULL DEFAULT ''")

        for key, value in DEFAULT_SETTINGS.items():
            conn.execute(
                "INSERT OR IGNORE INTO settings(key, value) VALUES(?, ?)",
                (key, value),
            )

        owner_id = get_setting_from_conn(conn, "owner_admin_id", "").strip()
        if not owner_id:
            oldest_admin = conn.execute(
                "SELECT telegram_user_id FROM admin_users ORDER BY added_at ASC, telegram_user_id ASC LIMIT 1"
            ).fetchone()
            if oldest_admin:
                owner_id = str(oldest_admin["telegram_user_id"])
                set_setting_with_conn(conn, "owner_admin_id", owner_id)
                conn.execute(
                    "UPDATE admin_users SET role='owner' WHERE telegram_user_id=?",
                    (int(owner_id),),
                )

        conn.commit()


def ensure_owner_from_env() -> int | None:
    raw = OWNER_ADMIN_ID.strip()
    if not raw:
        logger.warning("OWNER_ADMIN_ID is not set; automatic owner recovery is disabled.")
        return None

    if not raw.isdigit():
        logger.error("OWNER_ADMIN_ID must be numeric. Current value: %r", raw)
        return None

    owner_user_id = int(raw)
    with closing(get_db()) as conn:
        conn.execute(
            """
            INSERT INTO admin_users(telegram_user_id, telegram_username, role, added_at, added_by)
            VALUES(?, COALESCE((SELECT telegram_username FROM admin_users WHERE telegram_user_id=?), ''), 'owner', ?, ?)
            ON CONFLICT(telegram_user_id) DO UPDATE SET role='owner'
            """,
            (owner_user_id, owner_user_id, now_local().isoformat(), owner_user_id),
        )
        set_setting_with_conn(conn, "owner_admin_id", str(owner_user_id))
        conn.execute(
            "UPDATE admin_users SET role='admin' WHERE telegram_user_id<>? AND role='owner'",
            (owner_user_id,),
        )
        conn.commit()
    logger.info("Owner ensured from OWNER_ADMIN_ID=%s", owner_user_id)
    return owner_user_id


def ensure_default_admins_from_env(owner_user_id: int | None = None):
    admin_ids = parse_admin_ids(DEFAULT_ADMIN_IDS)
    if owner_user_id is not None:
        admin_ids = [user_id for user_id in admin_ids if user_id != owner_user_id]

    if not admin_ids:
        if DEFAULT_ADMIN_IDS.strip():
            logger.info("DEFAULT_ADMIN_IDS provided, but no additional valid admin ids were found.")
        return

    now_iso = now_local().isoformat()
    with closing(get_db()) as conn:
        for admin_user_id in admin_ids:
            conn.execute(
                """
                INSERT INTO admin_users(telegram_user_id, telegram_username, role, added_at, added_by)
                VALUES(?, COALESCE((SELECT telegram_username FROM admin_users WHERE telegram_user_id=?), ''), 'admin', ?, ?)
                ON CONFLICT(telegram_user_id) DO UPDATE SET role=CASE WHEN admin_users.role='owner' THEN 'owner' ELSE 'admin' END
                """,
                (admin_user_id, admin_user_id, now_iso, owner_user_id),
            )
        conn.commit()
    logger.info("Default admins ensured from DEFAULT_ADMIN_IDS=%s", ",".join(str(i) for i in admin_ids))




def parse_admin_ids(raw: str) -> list[int]:
    ids: list[int] = []
    seen: set[int] = set()
    for part in (raw or "").split(','):
        value = part.strip()
        if not value:
            continue
        if not value.isdigit():
            logger.warning("Ignoring non-numeric admin id in DEFAULT_ADMIN_IDS: %r", value)
            continue
        user_id = int(value)
        if user_id not in seen:
            ids.append(user_id)
            seen.add(user_id)
    return ids

def get_setting(key: str, default: str = "") -> str:
    with closing(get_db()) as conn:
        return get_setting_from_conn(conn, key, default)


def set_setting(key: str, value: str):
    with closing(get_db()) as conn:
        set_setting_with_conn(conn, key, value)
        conn.commit()


def now_local() -> datetime:
    return datetime.now(TZ)


def today_iso() -> str:
    return now_local().date().isoformat()


def yesterday_iso() -> str:
    return (now_local().date() - timedelta(days=1)).isoformat()


def plus_27_hours_iso() -> str:
    return (now_local() + timedelta(hours=27)).isoformat()


def fmt_dt_local(iso_value: str | None) -> str:
    if not iso_value:
        return "N/A"
    try:
        return datetime.fromisoformat(iso_value).astimezone(TZ).strftime("%Y-%m-%d %I:%M:%S %p")
    except Exception:
        return iso_value


def get_limits():
    return {
        "modems_total": int(get_setting("limit_modems_total", "12")),
        "dvr_total": int(get_setting("limit_dvr_total", "5")),
        "xg2": int(get_setting("limit_xg2", "5")),
        "xid": int(get_setting("limit_xid", "12")),
        "xi6": int(get_setting("limit_xi6", "12")),
        "xer10": int(get_setting("limit_xer10", "2")),
        "onu": int(get_setting("limit_onu", "2")),
        "screen": int(get_setting("limit_screen", "5")),
        "battery": int(get_setting("limit_battery", "50")),
        "sensor": int(get_setting("limit_sensor", "50")),
        "camera": int(get_setting("limit_camera", "50")),
        "extra_item": int(get_setting("limit_extra_item", "10")),
    }


def get_owner_admin_id() -> int | None:
    raw = get_setting("owner_admin_id", "").strip()
    if raw.isdigit():
        return int(raw)
    return None


def is_owner(user_id: int) -> bool:
    owner_id = get_owner_admin_id()
    return owner_id == user_id if owner_id is not None else False


def add_admin_user(user_id: int, username: str = "", role: str = "admin", added_by: int | None = None):
    now = now_local().isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            """
            INSERT INTO admin_users(telegram_user_id, telegram_username, role, added_at, added_by)
            VALUES(?, ?, ?, ?, ?)
            ON CONFLICT(telegram_user_id) DO UPDATE SET
                telegram_username=excluded.telegram_username,
                role=excluded.role
            """,
            (user_id, username or "", role, now, added_by),
        )
        if role == "owner":
            set_setting_with_conn(conn, "owner_admin_id", str(user_id))
        conn.commit()


def remove_admin_user(user_id: int) -> bool:
    owner_id = get_owner_admin_id()
    if owner_id == user_id:
        return False

    with closing(get_db()) as conn:
        cur = conn.execute("DELETE FROM admin_users WHERE telegram_user_id=?", (user_id,))
        conn.commit()
        return (cur.rowcount or 0) > 0


def is_admin(user_id: int) -> bool:
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT telegram_user_id FROM admin_users WHERE telegram_user_id=?",
            (user_id,),
        ).fetchone()
        return row is not None


def list_admin_users():
    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT telegram_user_id, telegram_username, role, added_at, added_by
            FROM admin_users
            ORDER BY CASE WHEN role='owner' THEN 0 ELSE 1 END, added_at ASC, telegram_user_id ASC
            """
        ).fetchall()
        return [dict(r) for r in rows]


def parse_int(value, default=0):
    try:
        return max(0, int(str(value).strip() or default))
    except Exception:
        return default


def get_active_message_info():
    message = get_setting("technician_message", "").strip()
    active_until = get_setting("technician_message_active_until", "").strip()

    if not message or not active_until:
        return "", ""

    try:
        expires_at = datetime.fromisoformat(active_until)
    except Exception:
        clear_technician_message()
        return "", ""

    if now_local() > expires_at:
        clear_technician_message()
        return "", ""

    return message, active_until


def get_active_technician_message() -> str:
    message, _ = get_active_message_info()
    return message


def set_technician_message(message: str):
    set_setting("technician_message", message.strip())
    set_setting("technician_message_active_until", plus_27_hours_iso())


def clear_technician_message():
    set_setting("technician_message", "")
    set_setting("technician_message_active_until", "")


def normalize_basic_key(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())
    return cleaned


def normalize_tech_id(value: str) -> str:
    return normalize_basic_key(value)


def normalize_bp_number(value: str) -> str:
    cleaned = normalize_basic_key(value)
    if cleaned.startswith("bp") and len(cleaned) > 2:
        cleaned = cleaned[2:]
    return cleaned


def combine_extra_item_names(existing_name: str, new_name: str) -> str:
    left = (existing_name or "").strip()
    right = (new_name or "").strip()

    if not left:
        return right
    if not right:
        return left
    if normalize_basic_key(left) == normalize_basic_key(right):
        return left

    return f"{left} | {right}"


def combine_notes(existing_notes: str, new_notes: str) -> str:
    left = (existing_notes or "").strip()
    right = (new_notes or "").strip()

    if not left:
        return right
    if not right:
        return left
    if left == right:
        return left

    return f"{left}\n{right}"


def merge_payload_with_existing(existing_row: dict, payload: dict) -> dict:
    merged = dict(payload)

    merged["telegram_user_id"] = payload.get("telegram_user_id") or existing_row.get("telegram_user_id")
    merged["telegram_username"] = (payload.get("telegram_username") or "").strip() or (existing_row.get("telegram_username") or "").strip()
    merged["tech_id"] = (existing_row.get("tech_id") or payload.get("tech_id") or "").strip()
    merged["bp_number"] = (existing_row.get("bp_number") or payload.get("bp_number") or "").strip()

    for field in MERGE_SUM_FIELDS:
        merged[field] = int(existing_row.get(field, 0) or 0) + int(payload.get(field, 0) or 0)

    merged["extra_item_name"] = combine_extra_item_names(
        existing_row.get("extra_item_name", ""),
        payload.get("extra_item_name", ""),
    )
    merged["notes"] = combine_notes(
        existing_row.get("notes", ""),
        payload.get("notes", ""),
    )

    return merged


def save_order(payload: dict):
    now = now_local()
    with closing(get_db()) as conn:
        conn.execute(
            """
            INSERT INTO orders (
                created_at, created_date, telegram_user_id, telegram_username,
                tech_id, bp_number,
                xb3, xb6, xb7, xb8, xb10,
                xg1, xg1_4k, xg2, xid, xi6, xer10, onu,
                screen, battery, sensor, camera,
                extra_item_name, extra_item_qty, notes, status, exported_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now.isoformat(),
                now.date().isoformat(),
                payload.get("telegram_user_id"),
                payload.get("telegram_username", ""),
                payload["tech_id"],
                payload["bp_number"],
                payload.get("xb3", 0),
                payload.get("xb6", 0),
                payload.get("xb7", 0),
                payload.get("xb8", 0),
                payload.get("xb10", 0),
                payload.get("xg1", 0),
                payload.get("xg1_4k", 0),
                payload.get("xg2", 0),
                payload.get("xid", 0),
                payload.get("xi6", 0),
                payload.get("xer10", 0),
                payload.get("onu", 0),
                payload.get("screen", 0),
                payload.get("battery", 0),
                payload.get("sensor", 0),
                payload.get("camera", 0),
                payload.get("extra_item_name", ""),
                payload.get("extra_item_qty", 0),
                payload.get("notes", ""),
                "active",
                "",
            ),
        )
        conn.commit()


def update_order(order_id: int, payload: dict):
    with closing(get_db()) as conn:
        conn.execute(
            """
            UPDATE orders SET
                telegram_user_id=?,
                telegram_username=?,
                tech_id=?,
                bp_number=?,
                xb3=?,
                xb6=?,
                xb7=?,
                xb8=?,
                xb10=?,
                xg1=?,
                xg1_4k=?,
                xg2=?,
                xid=?,
                xi6=?,
                xer10=?,
                onu=?,
                screen=?,
                battery=?,
                sensor=?,
                camera=?,
                extra_item_name=?,
                extra_item_qty=?,
                notes=?
            WHERE id=?
            """,
            (
                payload.get("telegram_user_id"),
                payload.get("telegram_username", ""),
                payload["tech_id"],
                payload["bp_number"],
                payload.get("xb3", 0),
                payload.get("xb6", 0),
                payload.get("xb7", 0),
                payload.get("xb8", 0),
                payload.get("xb10", 0),
                payload.get("xg1", 0),
                payload.get("xg1_4k", 0),
                payload.get("xg2", 0),
                payload.get("xid", 0),
                payload.get("xi6", 0),
                payload.get("xer10", 0),
                payload.get("onu", 0),
                payload.get("screen", 0),
                payload.get("battery", 0),
                payload.get("sensor", 0),
                payload.get("camera", 0),
                payload.get("extra_item_name", ""),
                payload.get("extra_item_qty", 0),
                payload.get("notes", ""),
                order_id,
            ),
        )
        conn.commit()


def find_matching_active_order(day_iso: str, tech_id: str, bp_number: str) -> dict | None:
    tech_key = normalize_tech_id(tech_id)
    bp_key = normalize_bp_number(bp_number)

    if not tech_key or not bp_key:
        return None

    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT * FROM orders
            WHERE created_date=? AND status='active'
            ORDER BY created_at DESC, id DESC
            """,
            (day_iso,),
        ).fetchall()

    for row in rows:
        row_dict = dict(row)
        if (
            normalize_tech_id(row_dict.get("tech_id", "")) == tech_key
            and normalize_bp_number(row_dict.get("bp_number", "")) == bp_key
        ):
            return row_dict
    return None


def save_or_merge_order(payload: dict):
    day_iso = today_iso()
    existing = find_matching_active_order(day_iso, payload["tech_id"], payload["bp_number"])

    if not existing:
        save_order(payload)
        return {"merged": False, "message": "Your request was sent to WH successfully."}

    merged_payload = merge_payload_with_existing(existing, payload)
    error = validate_payload(merged_payload)
    if error:
        return {"merged": True, "error": error}

    update_order(int(existing["id"]), merged_payload)
    return {"merged": True, "message": "Your request was added to the existing active order successfully."}


def fetch_orders_for_day(day_iso: str, status: str = "active"):
    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT * FROM orders
            WHERE created_date=? AND status=?
            ORDER BY created_at ASC, id ASC
            """,
            (day_iso, status),
        ).fetchall()
        return [dict(r) for r in rows]


def fetch_orders_since(start_date_iso: str, status: str = "active"):
    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT * FROM orders
            WHERE created_date>=? AND status=?
            ORDER BY created_at DESC, id DESC
            """,
            (start_date_iso, status),
        ).fetchall()
        return [dict(r) for r in rows]


def fetch_active_cycle_orders(status: str = "active"):
    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT * FROM orders
            WHERE status=?
            ORDER BY created_at ASC, id ASC
            """,
            (status,),
        ).fetchall()
        return [dict(r) for r in rows]


def fetch_active_cycle_orders_until(cutoff_iso: str, status: str = "active"):
    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT * FROM orders
            WHERE status=? AND created_at<=?
            ORDER BY created_at ASC, id ASC
            """,
            (status, cutoff_iso),
        ).fetchall()
        return [dict(r) for r in rows]


def get_active_cycle_start_iso() -> str:
    with closing(get_db()) as conn:
        row = conn.execute(
            """
            SELECT created_at
            FROM orders
            WHERE status='active'
            ORDER BY created_at ASC, id ASC
            LIMIT 1
            """
        ).fetchone()

    return row["created_at"] if row else ""


def enumerate_orders(rows: list[dict]) -> list[dict]:
    numbered = []
    for index, row in enumerate(rows, start=1):
        item = dict(row)
        item["daily_number"] = index
        numbered.append(item)
    return numbered


def delete_orders_by_ids(order_ids: list[int]) -> int:
    if not order_ids:
        return 0

    placeholders = ",".join(["?"] * len(order_ids))
    with closing(get_db()) as conn:
        cur = conn.execute(f"DELETE FROM orders WHERE id IN ({placeholders})", tuple(order_ids))
        conn.commit()
        return cur.rowcount or 0


def move_orders_to_history(order_ids: list[int], exported_at_iso: str | None = None) -> int:
    if not order_ids:
        return 0

    exported_at = exported_at_iso or now_local().isoformat()
    placeholders = ",".join(["?"] * len(order_ids))
    with closing(get_db()) as conn:
        cur = conn.execute(
            f"""
            UPDATE orders
            SET status='history', exported_at=?
            WHERE id IN ({placeholders}) AND status='active'
            """,
            (exported_at, *order_ids),
        )
        conn.commit()
        return cur.rowcount or 0


def get_cycle_order_by_number(order_number: int, status: str = "active") -> dict | None:
    rows = enumerate_orders(fetch_active_cycle_orders(status=status))
    for row in rows:
        if row["daily_number"] == order_number:
            return row
    return None


def build_csv_bytes_from_rows(rows: list[dict]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "order_number",
        "created_at",
        "created_date",
        "telegram_user_id",
        "telegram_username",
        "tech_id",
        "bp_number",
        "xb3",
        "xb6",
        "xb7",
        "xb8",
        "xb10",
        "xg1",
        "xg1_4k",
        "xg2",
        "xid",
        "xi6",
        "xer10",
        "onu",
        "screen",
        "battery",
        "sensor",
        "camera",
        "extra_item_name",
        "extra_item_qty",
        "notes",
        "status",
        "exported_at",
    ])
    for row in enumerate_orders(rows):
        writer.writerow([
            row["daily_number"],
            row["created_at"],
            row["created_date"],
            row["telegram_user_id"],
            row["telegram_username"],
            row["tech_id"],
            row["bp_number"],
            row["xb3"],
            row["xb6"],
            row["xb7"],
            row["xb8"],
            row["xb10"],
            row["xg1"],
            row["xg1_4k"],
            row["xg2"],
            row["xid"],
            row["xi6"],
            row["xer10"],
            row["onu"],
            row["screen"],
            row["battery"],
            row["sensor"],
            row["camera"],
            row["extra_item_name"],
            row["extra_item_qty"],
            row["notes"],
            row.get("status", "active"),
            row.get("exported_at", ""),
        ])
    return output.getvalue().encode("utf-8-sig")


def validate_payload(payload: dict):
    limits = get_limits()

    tech_id = payload["tech_id"].strip()
    bp_number = payload["bp_number"].strip()

    if not tech_id:
        return "Tech ID is required."
    if not bp_number:
        return "BP Number is required."

    modem_total = payload["xb3"] + payload["xb6"] + payload["xb7"] + payload["xb8"] + payload["xb10"]
    dvr_total = payload["xg1"] + payload["xg1_4k"]

    if modem_total > limits["modems_total"]:
        return f"Modem category limit reached. Maximum allowed is {limits['modems_total']}."
    if dvr_total > limits["dvr_total"]:
        return f"DVR category limit reached. Maximum allowed is {limits['dvr_total']}."
    if payload["xg2"] > limits["xg2"]:
        return f"XG2 limit reached. Maximum allowed is {limits['xg2']}."
    if payload["xid"] > limits["xid"]:
        return f"XID limit reached. Maximum allowed is {limits['xid']}."
    if payload["xi6"] > limits["xi6"]:
        return f"XI6 limit reached. Maximum allowed is {limits['xi6']}."
    if payload["xer10"] > limits["xer10"]:
        return f"XER10 limit reached. Maximum allowed is {limits['xer10']}."
    if payload["onu"] > limits["onu"]:
        return f"ONU limit reached. Maximum allowed is {limits['onu']}."
    if payload["screen"] > limits["screen"]:
        return f"Screen limit reached. Maximum allowed is {limits['screen']}."
    if payload["battery"] > limits["battery"]:
        return f"Battery limit reached. Maximum allowed is {limits['battery']}."
    if payload["sensor"] > limits["sensor"]:
        return f"Sensor limit reached. Maximum allowed is {limits['sensor']}."
    if payload["camera"] > limits["camera"]:
        return f"Camera limit reached. Maximum allowed is {limits['camera']}."

    if payload["extra_item_qty"] > 0 and not payload["extra_item_name"].strip():
        return "Please enter the item name for Add Item."

    if payload["extra_item_qty"] > limits["extra_item"]:
        return f"Additional item limit reached. Maximum allowed is {limits['extra_item']}."

    grand_total = (
        modem_total
        + dvr_total
        + payload["xg2"]
        + payload["xid"]
        + payload["xi6"]
        + payload["xer10"]
        + payload["onu"]
        + payload["screen"]
        + payload["battery"]
        + payload["sensor"]
        + payload["camera"]
        + payload["extra_item_qty"]
    )
    if grand_total <= 0:
        return "Please add at least one equipment item."

    return None


# ============================================================
# TECH BOT
# ============================================================
async def tech_start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    keyboard = [[
        InlineKeyboardButton(
            text="Open Order App",
            web_app=WebAppInfo(
                url=f"{BASE_URL}/webapp?uid={user.id}&username={user.username or ''}"
            ),
        )
    ]]
    await update.message.reply_text(
        "Open the app to submit equipment requests.",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


def register_tech_handlers():
    global _tech_handlers_registered
    if tech_bot_app is None or _tech_handlers_registered:
        return
    tech_bot_app.add_handler(CommandHandler("start", tech_start_command))
    _tech_handlers_registered = True


def run_tech_bot():
    if tech_bot_app is None:
        logger.info("Tech bot disabled.")
        return
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        register_tech_handlers()
        logger.info("Tech bot started.")
        tech_bot_app.run_polling(
            close_loop=False,
            drop_pending_updates=True,
            stop_signals=None,
        )
    except Exception as exc:
        logger.exception("Tech bot failed: %s", exc)


# ============================================================
# ADMIN BOT
# ============================================================
SETTABLE_LIMITS = {
    "limit_modems_total": "Modems Total",
    "limit_dvr_total": "DVR Total",
    "limit_xg2": "XG2",
    "limit_xid": "XID",
    "limit_xi6": "XI6",
    "limit_xer10": "XER10",
    "limit_onu": "ONU",
    "limit_screen": "Screen",
    "limit_battery": "Battery",
    "limit_sensor": "Sensor",
    "limit_camera": "Camera",
    "limit_extra_item": "Additional Item",
}


def equipment_totals(rows):
    totals = {k: 0 for k in EQUIPMENT_ORDER}
    for row in rows:
        for k in EQUIPMENT_ORDER:
            totals[k] += int(row.get(k, 0) or 0)
    return totals


def format_totals_multiline(totals: dict) -> str:
    lines = []
    for key in EQUIPMENT_ORDER:
        value = int(totals.get(key, 0) or 0)
        if value > 0:
            lines.append(f"- {EQUIPMENT_LABELS[key]}: {value}")
    return "\n".join(lines) if lines else "- No equipment requested"


def build_export_filename(is_excel: bool = True) -> str:
    stamp = now_local().strftime("%Y-%m-%d_%I-%M-%S_%p")
    ext = "xlsx" if is_excel else "csv"
    return f"orders_export_{stamp}.{ext}"


def admin_home_keyboard():
    rows = [
        [KeyboardButton("Start"), KeyboardButton("View Orders")],
        [KeyboardButton("Order History"), KeyboardButton("Delete Order")],
        [KeyboardButton("Export Orders"), KeyboardButton("View Statistics")],
        [KeyboardButton("Message for Technicians"), KeyboardButton("Set Max Equipment")],
        [KeyboardButton("Manage Admins")],
    ]

    return ReplyKeyboardMarkup(
        keyboard=rows,
        resize_keyboard=True,
        one_time_keyboard=False,
        is_persistent=True,
    )


def export_orders_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Export and Keep Orders", callback_data="export_keep")],
        [InlineKeyboardButton("Export and Delete Exported Orders", callback_data="export_delete")],
        [InlineKeyboardButton("Back", callback_data="back_main")],
    ])


def admin_message_menu():
    message, _ = get_active_message_info()
    rows = []
    if message:
        rows.append([InlineKeyboardButton("Edit Message", callback_data="edit_message")])
        rows.append([InlineKeyboardButton("Delete Message", callback_data="delete_message")])
    else:
        rows.append([InlineKeyboardButton("Create Message", callback_data="create_message")])

    rows.append([InlineKeyboardButton("Back", callback_data="back_main")])
    return InlineKeyboardMarkup(rows)


def admin_limits_menu():
    rows = []
    current_row = []
    for key, label in SETTABLE_LIMITS.items():
        current_value = get_setting(key, "0")
        current_row.append(
            InlineKeyboardButton(f"{label}: {current_value}", callback_data=f"limit::{key}")
        )
        if len(current_row) == 2:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)
    rows.append([InlineKeyboardButton("Back", callback_data="back_main")])
    return InlineKeyboardMarkup(rows)


def manage_admins_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("List Admins", callback_data="list_admins")],
        [InlineKeyboardButton("Add Admin", callback_data="add_admin_prompt")],
        [InlineKeyboardButton("Remove Admin", callback_data="remove_admin_prompt")],
        [InlineKeyboardButton("Back", callback_data="back_main")],
    ])


def admin_message_status_text():
    message, active_until = get_active_message_info()
    if not message:
        return "No active technician message."

    return (
        "Current technician message:\n\n"
        f"{message}\n\n"
        f"Active until: {fmt_dt_local(active_until)}"
    )


def weekly_stats_text():
    start_date = (now_local().date() - timedelta(days=6)).isoformat()
    rows = fetch_orders_since(start_date, status="active") + fetch_orders_since(start_date, status="history")

    totals = equipment_totals(rows)

    tech_visits = {}
    for row in rows:
        tech = row.get("tech_id") or "Unknown"
        tech_visits[tech] = tech_visits.get(tech, 0) + 1

    top_equipment = sorted(totals.items(), key=lambda x: x[1], reverse=True)
    top_techs = sorted(tech_visits.items(), key=lambda x: x[1], reverse=True)

    lines = [
        "📊 Weekly Statistics",
        f"Range: {start_date} to {today_iso()}",
        "",
        "Most requested equipment:",
    ]

    shown_any = False
    for key, value in top_equipment:
        if value > 0:
            lines.append(f"- {EQUIPMENT_LABELS.get(key, key.upper())}: {value}")
            shown_any = True
    if not shown_any:
        lines.append("- No orders this week.")

    lines.append("")
    lines.append("Technician WH visits this week:")
    if top_techs:
        for tech, visits in top_techs:
            lines.append(f"- {tech}: {visits}")
    else:
        lines.append("- No visits recorded.")

    return "\n".join(lines)


def _items_for_row(row: dict) -> str:
    items = []
    for key in EQUIPMENT_ORDER[:-1]:
        qty = int(row.get(key, 0) or 0)
        if qty > 0:
            items.append(f"{EQUIPMENT_LABELS[key]} {qty}")

    extra_name = (row.get("extra_item_name") or "").strip()
    extra_qty = int(row.get("extra_item_qty", 0) or 0)
    if extra_qty > 0:
        items.append(f"{extra_name or 'Additional Item'} {extra_qty}")

    return ", ".join(items) if items else "No items"


def current_cycle_orders_text():
    cycle_start = get_active_cycle_start_iso()
    rows = enumerate_orders(fetch_active_cycle_orders(status="active"))

    if not rows:
        return (
            "📦 Current Active Orders\n"
            "Cycle start: waiting for first new order\n\n"
            "No active orders found in the current cycle."
        )

    totals = equipment_totals(rows)
    lines = [
        "📦 Current Active Orders",
        f"Cycle start: {fmt_dt_local(cycle_start)}",
        f"Total orders: {len(rows)}",
        "",
        "Equipment totals:",
        format_totals_multiline(totals),
        "",
        "Order details:",
    ]

    for row in rows:
        created_display = fmt_dt_local(row.get("created_at"))
        lines.append(
            f"#{row['daily_number']} | {created_display} | Tech {row['tech_id']} | BP {row['bp_number']} | {_items_for_row(row)}"
        )

    return "\n".join(lines)


def order_history_text():
    start_date = (now_local().date() - timedelta(days=13)).isoformat()
    rows = fetch_orders_since(start_date, status="history")

    if not rows:
        return "🗂️ Order History\n\nNo archived orders found in the last 14 days."

    grouped = {}
    for row in rows:
        export_day = "Unknown"
        if row.get("exported_at"):
            try:
                export_day = datetime.fromisoformat(row["exported_at"]).astimezone(TZ).date().isoformat()
            except Exception:
                export_day = "Unknown"
        grouped.setdefault(export_day, []).append(row)

    lines = ["🗂️ Order History", "Showing archived orders from the last 14 days.", ""]

    for day_iso in sorted(grouped.keys(), reverse=True):
        daily_rows = list(sorted(grouped[day_iso], key=lambda r: (r.get("exported_at", ""), r["created_at"], r["id"])))
        numbered_rows = enumerate_orders(daily_rows)
        totals = equipment_totals(numbered_rows)

        lines.append(f"Export date: {day_iso}")
        lines.append(f"Total archived orders: {len(numbered_rows)}")
        lines.append("Equipment totals:")
        lines.append(format_totals_multiline(totals))
        lines.append("Order details:")

        for row in numbered_rows:
            lines.append(
                f"#{row['daily_number']} | {fmt_dt_local(row.get('created_at'))} | "
                f"Exported {fmt_dt_local(row.get('exported_at'))} | "
                f"Tech {row['tech_id']} | BP {row['bp_number']} | {_items_for_row(row)}"
            )

        lines.append("")

    return "\n".join(lines).strip()


def list_admins_text():
    rows = list_admin_users()
    if not rows:
        return "No admins configured."

    lines = ["👮 Admin Users", ""]
    for row in rows:
        username = f"@{row['telegram_username']}" if row.get("telegram_username") else "(no username)"
        lines.append(
            f"- {row['role'].upper()} | ID {row['telegram_user_id']} | {username}"
        )
    lines.append("")
    lines.append("Tip: any user can send /myid to the bot to see their Telegram ID.")
    return "\n".join(lines)


def build_excel_summary_from_rows(cycle_start_iso: str, export_time_iso: str, rows: list[dict]) -> bytes:
    output = io.BytesIO()
    numbered_rows = enumerate_orders(rows)

    try:
        from openpyxl import Workbook
    except Exception:
        csv_bytes = build_csv_bytes_from_rows(rows)
        output.write(csv_bytes)
        output.seek(0)
        return output.getvalue()

    wb = Workbook()

    totals = equipment_totals(numbered_rows)

    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.append(["Cycle Start", fmt_dt_local(cycle_start_iso)])
    ws_overview.append(["Exported At", fmt_dt_local(export_time_iso)])
    ws_overview.append(["Total Orders", len(numbered_rows)])
    ws_overview.append([])
    ws_overview.append(["Equipment", "Total Qty"])
    for key in EQUIPMENT_ORDER:
        ws_overview.append([EQUIPMENT_LABELS[key], totals.get(key, 0)])

    ws_summary = wb.create_sheet("Summary by Tech")
    ws_summary.append([
        "Tech ID",
        "BP Number",
        "Orders",
        "XB3", "XB6", "XB7", "XB8", "XB10",
        "XG1", "XG1 4K", "XG2", "XID", "XI6",
        "XER10", "ONU", "Screen", "Battery", "Sensor", "Camera",
        "Extra Item", "Extra Qty",
    ])

    grouped = {}
    for row in numbered_rows:
        key = (row["tech_id"], row["bp_number"])
        if key not in grouped:
            grouped[key] = {
                "orders": 0,
                "xb3": 0, "xb6": 0, "xb7": 0, "xb8": 0, "xb10": 0,
                "xg1": 0, "xg1_4k": 0, "xg2": 0, "xid": 0, "xi6": 0,
                "xer10": 0, "onu": 0, "screen": 0, "battery": 0, "sensor": 0, "camera": 0,
                "extra_item_name": "",
                "extra_item_qty": 0,
            }

        grouped[key]["orders"] += 1
        for field in [
            "xb3", "xb6", "xb7", "xb8", "xb10",
            "xg1", "xg1_4k", "xg2", "xid", "xi6",
            "xer10", "onu", "screen", "battery", "sensor", "camera",
        ]:
            grouped[key][field] += int(row.get(field, 0) or 0)

        if row.get("extra_item_name"):
            grouped[key]["extra_item_name"] = row["extra_item_name"]
        grouped[key]["extra_item_qty"] += int(row.get("extra_item_qty", 0) or 0)

    for (tech_id, bp_number), data in grouped.items():
        ws_summary.append([
            tech_id,
            bp_number,
            data["orders"],
            data["xb3"], data["xb6"], data["xb7"], data["xb8"], data["xb10"],
            data["xg1"], data["xg1_4k"], data["xg2"], data["xid"], data["xi6"],
            data["xer10"], data["onu"], data["screen"], data["battery"], data["sensor"], data["camera"],
            data["extra_item_name"], data["extra_item_qty"],
        ])

    ws_raw = wb.create_sheet("Raw Orders")
    ws_raw.append([
        "order_number",
        "created_at",
        "created_date",
        "tech_id",
        "bp_number",
        "xb3", "xb6", "xb7", "xb8", "xb10",
        "xg1", "xg1_4k", "xg2", "xid", "xi6",
        "xer10", "onu", "screen", "battery", "sensor", "camera",
        "extra_item_name", "extra_item_qty", "notes", "status", "exported_at"
    ])
    for row in numbered_rows:
        ws_raw.append([
            row["daily_number"],
            row["created_at"],
            row["created_date"],
            row["tech_id"],
            row["bp_number"],
            row["xb3"], row["xb6"], row["xb7"], row["xb8"], row["xb10"],
            row["xg1"], row["xg1_4k"], row["xg2"], row["xid"], row["xi6"],
            row["xer10"], row["onu"], row["screen"], row["battery"], row["sensor"], row["camera"],
            row["extra_item_name"], row["extra_item_qty"], row["notes"], row.get("status", "active"), row.get("exported_at", "")
        ])

    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def send_orders_export(chat_id: int, bot, cycle_start_iso: str, export_time_iso: str, rows: list[dict], delete_after_export: bool):
    totals = equipment_totals(rows)
    file_bytes = build_excel_summary_from_rows(cycle_start_iso, export_time_iso, rows)
    is_excel = file_bytes[:2] == b"PK"
    filename = build_export_filename(is_excel=is_excel)

    caption_lines = [
        "Export completed",
        f"Cycle start: {fmt_dt_local(cycle_start_iso)}",
        f"Exported at: {fmt_dt_local(export_time_iso)}",
        f"Total orders: {len(rows)}",
        "",
        "Equipment totals:",
        format_totals_multiline(totals),
        "",
        f"Mode: {'Export and Delete Exported Orders' if delete_after_export else 'Export and Keep Orders'}",
        "Archived to Order History: Yes",
    ]

    await bot.send_document(
        chat_id=chat_id,
        document=InputFile(io.BytesIO(file_bytes), filename=filename),
        caption="\n".join(caption_lines),
    )


async def run_export_action(chat_id: int, bot, delete_after_export: bool):
    export_time_iso = now_local().isoformat()
    rows = fetch_active_cycle_orders_until(export_time_iso, status="active")

    if not rows:
        await bot.send_message(
            chat_id=chat_id,
            text="No active orders found in the current cycle.",
        )
        return

    cycle_start_iso = rows[0]["created_at"]

    await send_orders_export(
        chat_id=chat_id,
        bot=bot,
        cycle_start_iso=cycle_start_iso,
        export_time_iso=export_time_iso,
        rows=rows,
        delete_after_export=delete_after_export,
    )

    order_ids = [int(r["id"]) for r in rows]
    moved_count = move_orders_to_history(order_ids, exported_at_iso=export_time_iso)

    await bot.send_message(
        chat_id=chat_id,
        text=(
            "Export completed.\n"
            f"Archived exported orders to Order History: {moved_count}\n"
            "The next new order will start a new cycle."
        ),
    )


async def admin_safe_answer(query):
    try:
        await query.answer()
    except BadRequest as exc:
        if "Query is too old" in str(exc) or "query id is invalid" in str(exc):
            logger.warning("Ignored expired admin callback query.")
            return
        raise


async def admin_safe_edit_message(query, text, reply_markup=None):
    try:
        await query.edit_message_text(text, reply_markup=reply_markup)
    except BadRequest as exc:
        if "Message is not modified" in str(exc):
            return
        raise


async def admin_error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.exception("Unhandled admin bot error: %s", context.error)


async def admin_start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    owner_id = get_owner_admin_id()

    if owner_id is None:
        add_admin_user(user.id, user.username or "", role="owner", added_by=user.id)
        await update.message.reply_text(
            "Owner admin access granted automatically.",
            reply_markup=admin_home_keyboard(),
        )
        return

    if is_admin(user.id):
        add_admin_user(user.id, user.username or "", "owner" if is_owner(user.id) else "admin", None)
        await update.message.reply_text(
            "Admin bot ready. Use the buttons below.",
            reply_markup=admin_home_keyboard(),
        )
        return

    await update.message.reply_text(
        "This bot is restricted.\n"
        "Ask the owner to add you as admin.\n\n"
        "You can send /myid to share your Telegram ID.",
        reply_markup=admin_home_keyboard(),
    )


async def admin_myid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    username = f"@{user.username}" if user.username else "(no username)"
    await update.message.reply_text(
        f"Your Telegram ID is:\n{user.id}\n\nUsername: {username}"
    )


async def admin_authorize_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    if is_admin(user.id):
        await update.message.reply_text(
            "Admin access already active.",
            reply_markup=admin_home_keyboard(),
        )
        return

    await update.message.reply_text(
        "Manual token authorization is disabled.\n"
        "Ask the owner to add you from Manage Admins.\n\n"
        "Use /myid to get your Telegram ID."
    )


async def admin_callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await admin_safe_answer(query)

    user = query.from_user
    if not is_admin(user.id):
        await admin_safe_edit_message(query, "Unauthorized.")
        return

    data = query.data

    if data == "back_main":
        context.user_data.pop("awaiting", None)
        await admin_safe_edit_message(query, "Use the buttons below.")
        return

    if data == "view_orders":
        await admin_safe_edit_message(query, current_cycle_orders_text())
        return

    if data == "order_history":
        await admin_safe_edit_message(query, order_history_text())
        return

    if data == "delete_order_prompt":
        context.user_data["awaiting"] = "delete_order_number"
        await admin_safe_edit_message(
            query,
            "Send the current cycle order number to delete.\nExample: 3\n\nUse View Orders to see the current order numbers.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Back", callback_data="back_main")]])
        )
        return

    if data == "view_stats":
        await admin_safe_edit_message(query, weekly_stats_text())
        return

    if data == "export_orders_menu":
        context.user_data.pop("awaiting", None)
        await admin_safe_edit_message(
            query,
            "Choose how to export the current active cycle orders.",
            reply_markup=export_orders_menu(),
        )
        return

    if data == "export_keep":
        await run_export_action(query.message.chat_id, context.bot, delete_after_export=False)
        await admin_safe_edit_message(query, "Export completed. Use the buttons below.")
        return

    if data == "export_delete":
        await run_export_action(query.message.chat_id, context.bot, delete_after_export=True)
        await admin_safe_edit_message(query, "Export completed. Use the buttons below.")
        return

    if data == "message_menu":
        await admin_safe_edit_message(query, admin_message_status_text(), reply_markup=admin_message_menu())
        return

    if data == "create_message":
        context.user_data["awaiting"] = "technician_message_create"
        await admin_safe_edit_message(
            query,
            "Send the new message for technicians now.\nIt will expire automatically 27 hours after you send it.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Back", callback_data="message_menu")]])
        )
        return

    if data == "edit_message":
        context.user_data["awaiting"] = "technician_message_edit"
        await admin_safe_edit_message(
            query,
            "Send the updated message for technicians now.\nIt will expire automatically 27 hours after you send it.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Back", callback_data="message_menu")]])
        )
        return

    if data == "delete_message":
        clear_technician_message()
        context.user_data.pop("awaiting", None)
        await admin_safe_edit_message(query, "Technician message deleted.")
        return

    if data == "set_limits":
        await admin_safe_edit_message(query, "Select the equipment limit to update:", reply_markup=admin_limits_menu())
        return

    if data.startswith("limit::"):
        setting_key = data.split("::", 1)[1]
        context.user_data["awaiting"] = f"limit::{setting_key}"
        label = SETTABLE_LIMITS.get(setting_key, setting_key)
        await admin_safe_edit_message(
            query,
            f"Send the new value for {label}.\nAllowed range: 0 to 50.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Back", callback_data="set_limits")]])
        )
        return

    if data == "manage_admins":
        if not is_owner(user.id):
            await admin_safe_edit_message(query, "Only the owner can manage admin users.")
            return
        context.user_data.pop("awaiting", None)
        await admin_safe_edit_message(query, "Admin management", reply_markup=manage_admins_menu())
        return

    if data == "list_admins":
        if not is_owner(user.id):
            await admin_safe_edit_message(query, "Only the owner can manage admin users.")
            return
        await admin_safe_edit_message(query, list_admins_text(), reply_markup=manage_admins_menu())
        return

    if data == "add_admin_prompt":
        if not is_owner(user.id):
            await admin_safe_edit_message(query, "Only the owner can manage admin users.")
            return
        context.user_data["awaiting"] = "add_admin_id"
        await admin_safe_edit_message(
            query,
            "Send the Telegram ID of the user to add as admin.\n\nTip: the user can send /myid to this bot and share the number with you.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Back", callback_data="manage_admins")]])
        )
        return

    if data == "remove_admin_prompt":
        if not is_owner(user.id):
            await admin_safe_edit_message(query, "Only the owner can manage admin users.")
            return
        context.user_data["awaiting"] = "remove_admin_id"
        await admin_safe_edit_message(
            query,
            "Send the Telegram ID of the admin to remove.\nThe owner cannot be removed.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Back", callback_data="manage_admins")]])
        )
        return


async def admin_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    text = (update.message.text or "").strip()

    if not is_admin(user.id):
        if text.lower() in {
            "start", "view orders", "order history", "delete order",
            "export orders", "view statistics", "manage admins",
            "message for technicians", "set max equipment"
        }:
            await update.message.reply_text(
                "Unauthorized.\n\nIf you should have access, press /start first or ask the owner to add you as admin."
            )
        return

    awaiting = context.user_data.get("awaiting")

    if text == "Start":
        context.user_data.pop("awaiting", None)
        await admin_start_command(update, context)
        return

    if text == "View Orders":
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(current_cycle_orders_text(), reply_markup=admin_home_keyboard())
        return

    if text == "Order History":
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(order_history_text(), reply_markup=admin_home_keyboard())
        return

    if text == "Delete Order":
        context.user_data["awaiting"] = "delete_order_number"
        await update.message.reply_text(
            "Send the current cycle order number to delete.\nExample: 3\n\nUse View Orders to see the current order numbers.",
            reply_markup=admin_home_keyboard(),
        )
        return

    if text == "Export Orders":
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(
            "Choose how to export the current active cycle orders.",
            reply_markup=export_orders_menu(),
        )
        return

    if text == "View Statistics":
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(weekly_stats_text(), reply_markup=admin_home_keyboard())
        return

    if text == "Manage Admins":
        context.user_data.pop("awaiting", None)
        if not is_owner(user.id):
            await update.message.reply_text("Only the owner can manage admin users.", reply_markup=admin_home_keyboard())
            return
        await update.message.reply_text("Admin management", reply_markup=manage_admins_menu())
        return

    if text == "Message for Technicians":
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(admin_message_status_text(), reply_markup=admin_message_menu())
        return

    if text == "Set Max Equipment":
        context.user_data.pop("awaiting", None)
        await update.message.reply_text("Select the equipment limit to update:", reply_markup=admin_limits_menu())
        return

    if not awaiting:
        await update.message.reply_text("Use the buttons below.", reply_markup=admin_home_keyboard())
        return

    if awaiting in ("technician_message_create", "technician_message_edit"):
        set_technician_message(text)
        context.user_data.pop("awaiting", None)
        _, active_until = get_active_message_info()
        await update.message.reply_text(
            f"Technician message saved.\nActive until: {fmt_dt_local(active_until)}",
            reply_markup=admin_home_keyboard(),
        )
        return

    if awaiting.startswith("limit::"):
        try:
            value = int(text)
            if value < 0 or value > 50:
                raise ValueError
        except Exception:
            await update.message.reply_text("Enter a number from 0 to 50.")
            return

        setting_key = awaiting.split("::", 1)[1]
        set_setting(setting_key, str(value))
        context.user_data.pop("awaiting", None)
        label = SETTABLE_LIMITS.get(setting_key, setting_key)
        await update.message.reply_text(
            f"{label} updated to {value}.",
            reply_markup=admin_home_keyboard(),
        )
        return

    if awaiting == "delete_order_number":
        try:
            order_number = int(text)
            if order_number <= 0:
                raise ValueError
        except Exception:
            await update.message.reply_text("Enter a valid current cycle order number like 1, 2, or 3.")
            return

        row = get_cycle_order_by_number(order_number, status="active")
        if not row:
            await update.message.reply_text("That current cycle order number was not found.")
            return

        deleted_count = delete_orders_by_ids([int(row["id"])])
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(
            f"Order #{order_number} deleted.\nRemoved records: {deleted_count}",
            reply_markup=admin_home_keyboard(),
        )
        return

    if awaiting == "add_admin_id":
        if not is_owner(user.id):
            context.user_data.pop("awaiting", None)
            await update.message.reply_text("Only the owner can manage admin users.", reply_markup=admin_home_keyboard())
            return

        try:
            target_id = int(text)
            if target_id <= 0:
                raise ValueError
        except Exception:
            await update.message.reply_text("Send a valid numeric Telegram ID.")
            return

        add_admin_user(target_id, "", role="admin", added_by=user.id)
        context.user_data.pop("awaiting", None)
        await update.message.reply_text(
            f"Admin added successfully.\nTelegram ID: {target_id}",
            reply_markup=manage_admins_menu(),
        )
        return

    if awaiting == "remove_admin_id":
        if not is_owner(user.id):
            context.user_data.pop("awaiting", None)
            await update.message.reply_text("Only the owner can manage admin users.", reply_markup=admin_home_keyboard())
            return

        try:
            target_id = int(text)
            if target_id <= 0:
                raise ValueError
        except Exception:
            await update.message.reply_text("Send a valid numeric Telegram ID.")
            return

        if is_owner(target_id):
            await update.message.reply_text("The owner cannot be removed.")
            return

        removed = remove_admin_user(target_id)
        context.user_data.pop("awaiting", None)
        if removed:
            await update.message.reply_text(
                f"Admin removed successfully.\nTelegram ID: {target_id}",
                reply_markup=manage_admins_menu(),
            )
        else:
            await update.message.reply_text(
                "That Telegram ID is not currently an admin.",
                reply_markup=manage_admins_menu(),
            )
        return


async def admin_post_init(application: Application):
    await application.bot.set_my_commands([
        BotCommand("start", "Open admin menu"),
        BotCommand("myid", "Show your Telegram ID"),
    ])


def register_admin_handlers():
    global _admin_handlers_registered
    if admin_bot_app is None or _admin_handlers_registered:
        return

    admin_bot_app.post_init = admin_post_init
    admin_bot_app.add_handler(CommandHandler("start", admin_start_command))
    admin_bot_app.add_handler(CommandHandler("authorize", admin_authorize_command))
    admin_bot_app.add_handler(CommandHandler("myid", admin_myid_command))
    admin_bot_app.add_handler(CallbackQueryHandler(admin_callback_handler))
    admin_bot_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_text_handler))
    admin_bot_app.add_error_handler(admin_error_handler)
    _admin_handlers_registered = True


def run_admin_bot():
    if admin_bot_app is None:
        logger.info("Admin bot disabled.")
        return
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        register_admin_handlers()
        logger.info("Admin bot started.")
        admin_bot_app.run_polling(
            close_loop=False,
            drop_pending_updates=True,
            stop_signals=None,
        )
    except Exception as exc:
        logger.exception("Admin bot failed: %s", exc)


# ============================================================
# WEB ROUTES
# ============================================================
@app.get("/")
def root():
    return redirect("/webapp")


@app.get("/health")
def health():
    return jsonify({"ok": True})


@app.get("/webapp")
def webapp_page():
    limits = get_limits()
    technician_message = get_active_technician_message()
    popup_message = (
        "Badge is required to pick up equipment. "
        "The warehouse closes at 9:00 AM. "
        "Please check your buffer first. "
        f"Maximum allowed quantities: "
        f"Modems {limits['modems_total']} total, "
        f"XI6 {limits['xi6']}, "
        f"XID {limits['xid']}, "
        f"XG2 {limits['xg2']}, "
        f"DVR {limits['dvr_total']} total, "
        f"XER10 {limits['xer10']}, "
        f"ONU {limits['onu']}, "
        f"Screen {limits['screen']}, "
        f"Additional Item {limits['extra_item']}."
    )
    prefill = {
        "telegram_user_id": request.args.get("uid", "").strip(),
        "telegram_username": request.args.get("username", "").strip(),
    }
    return render_template(
        "request_form.html",
        limits=limits,
        popup_message=popup_message,
        technician_message=technician_message,
        prefill=prefill,
    )


@app.post("/create-order")
def create_order():
    payload = {
        "telegram_user_id": request.form.get("telegram_user_id"),
        "telegram_username": request.form.get("telegram_username", "").strip(),
        "tech_id": request.form.get("tech_id", "").strip(),
        "bp_number": request.form.get("bp_number", "").strip(),

        "xb3": parse_int(request.form.get("xb3")),
        "xb6": parse_int(request.form.get("xb6")),
        "xb7": parse_int(request.form.get("xb7")),
        "xb8": parse_int(request.form.get("xb8")),
        "xb10": parse_int(request.form.get("xb10")),

        "xg1": parse_int(request.form.get("xg1")),
        "xg1_4k": parse_int(request.form.get("xg1_4k")),
        "xg2": parse_int(request.form.get("xg2")),
        "xid": parse_int(request.form.get("xid")),
        "xi6": parse_int(request.form.get("xi6")),
        "xer10": parse_int(request.form.get("xer10")),
        "onu": parse_int(request.form.get("onu")),

        "screen": parse_int(request.form.get("screen")),
        "battery": parse_int(request.form.get("battery")),
        "sensor": parse_int(request.form.get("sensor")),
        "camera": parse_int(request.form.get("camera")),

        "extra_item_name": request.form.get("extra_item_name", "").strip(),
        "extra_item_qty": parse_int(request.form.get("extra_item_qty")),
        "notes": request.form.get("notes", "").strip(),
    }

    error = validate_payload(payload)
    if error:
        limits = get_limits()
        technician_message = get_active_technician_message()
        popup_message = (
            "Badge is required to pick up equipment. "
            "The warehouse closes at 9:00 AM. "
            "Please check your buffer first. "
            f"Maximum allowed quantities: "
            f"Modems {limits['modems_total']} total, "
            f"XI6 {limits['xi6']}, "
            f"XID {limits['xid']}, "
            f"XG2 {limits['xg2']}, "
            f"DVR {limits['dvr_total']} total, "
            f"XER10 {limits['xer10']}, "
            f"ONU {limits['onu']}, "
            f"Screen {limits['screen']}, "
            f"Additional Item {limits['extra_item']}."
        )
        return render_template(
            "request_form.html",
            limits=limits,
            popup_message=popup_message,
            technician_message=technician_message,
            form_data=payload,
            form_error=error,
            prefill={
                "telegram_user_id": payload.get("telegram_user_id", ""),
                "telegram_username": payload.get("telegram_username", ""),
            },
        ), 400

    result = save_or_merge_order(payload)
    if result.get("error"):
        limits = get_limits()
        technician_message = get_active_technician_message()
        popup_message = (
            "Badge is required to pick up equipment. "
            "The warehouse closes at 9:00 AM. "
            "Please check your buffer first. "
            f"Maximum allowed quantities: "
            f"Modems {limits['modems_total']} total, "
            f"XI6 {limits['xi6']}, "
            f"XID {limits['xid']}, "
            f"XG2 {limits['xg2']}, "
            f"DVR {limits['dvr_total']} total, "
            f"XER10 {limits['xer10']}, "
            f"ONU {limits['onu']}, "
            f"Screen {limits['screen']}, "
            f"Additional Item {limits['extra_item']}."
        )
        return render_template(
            "request_form.html",
            limits=limits,
            popup_message=popup_message,
            technician_message=technician_message,
            form_data=payload,
            form_error=result["error"],
            prefill={
                "telegram_user_id": payload.get("telegram_user_id", ""),
                "telegram_username": payload.get("telegram_username", ""),
            },
        ), 400

    return render_template(
        "success.html",
        confirmation_message=result["message"],
        bp_number=payload["bp_number"],
    )


@app.get("/admin/export")
def admin_export_api():
    token = request.args.get("token", "")
    if token != ADMIN_ACCESS_TOKEN:
        return jsonify({"error": "Unauthorized"}), 401

    date_param = (request.args.get("date") or "").strip()
    if date_param:
        rows = fetch_orders_for_day(date_param, status="active")
    else:
        rows = fetch_active_cycle_orders(status="active")

    csv_bytes = build_csv_bytes_from_rows(rows)
    filename = build_export_filename(is_excel=False)
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/admin/export/check")
def admin_export_check():
    date_param = (request.args.get("date") or "").strip()
    if date_param:
        rows = fetch_orders_for_day(date_param, status="active")
        label = date_param
    else:
        rows = fetch_active_cycle_orders(status="active")
        label = f"cycle_start={get_active_cycle_start_iso() or 'none'}"

    return jsonify({
        "ok": True,
        "scope": label,
        "orders_found": len(rows),
        "generated_at": now_local().isoformat(),
    })


@app.get("/healthz")
def healthz():
    try:
        with closing(get_db()) as conn:
            conn.execute("SELECT 1").fetchone()
            owner_id = get_setting_from_conn(conn, "owner_admin_id", "").strip()
        return jsonify({
            "ok": True,
            "db_path": DB_PATH,
            "owner_admin_id": owner_id,
            "run_tech_bot": RUN_TECH_BOT,
            "run_admin_bot": RUN_ADMIN_BOT,
            "time": now_local().isoformat(),
        }), 200
    except Exception as exc:
        logger.exception("Health check failed")
        return jsonify({"ok": False, "error": str(exc)}), 500


# ============================================================
# STARTUP
# ============================================================
_started = False
_start_lock = threading.Lock()


def ensure_app_started(start_bots: bool = False):
    global _started
    if _started:
        return
    with _start_lock:
        if _started:
            return

        init_db()
        owner_user_id = ensure_owner_from_env()
        ensure_default_admins_from_env(owner_user_id)

        should_run_tech = tech_bot_app is not None and (start_bots or RUN_TECH_BOT)
        should_run_admin = admin_bot_app is not None and (start_bots or RUN_ADMIN_BOT)

        if should_run_tech:
            tech_thread = threading.Thread(target=run_tech_bot, daemon=True, name="tech-bot-thread")
            tech_thread.start()
            logger.info("Tech bot thread started.")
        else:
            logger.info("Tech bot thread skipped.")

        if should_run_admin:
            admin_thread = threading.Thread(target=run_admin_bot, daemon=True, name="admin-bot-thread")
            admin_thread.start()
            logger.info("Admin bot thread started.")
        else:
            logger.info("Admin bot thread skipped.")

        _started = True
        logger.info("Application startup completed.")


@app.before_request
def _ensure_started_before_request():
    ensure_app_started(start_bots=False)


ensure_app_started(start_bots=False)


if __name__ == "__main__":
    ensure_app_started(start_bots=True)
    logger.info("Web app starting on %s:%s", APP_HOST, APP_PORT)
    logger.info("BASE_URL=%s", BASE_URL)
    app.run(host=APP_HOST, port=APP_PORT, debug=False, use_reloader=False)
